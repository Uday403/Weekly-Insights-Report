import os, glob
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook

# ================= CONFIG ================= #
FILE_PATH = None           # auto-detect latest file
SHEET_NAME = "Sheet 1"
CLEANED_SHEET = "Cleaned"
INSIGHTS_SHEET = "Insights"

COLS = {
    "dim_a": "Dimension 1",
    "dim_b": "Dimension 2",
    "campaign": "Dimension 3",
    "impressions": "Impressions",
    "clicks": "Clicks",
    "spend": "Cost",
    "conversions": "Sydney Conversions",
}
# =========================================== #

def ensure_file_path(fp=None):
    """Return explicit path if exists; else auto-pick latest report on Desktop.
       Handles both hyphen and en-dash OneDrive folder names."""
    if fp and os.path.exists(fp):
        return fp

    # Try both OneDrive folder spellings (hyphen and en-dash)
    base = os.path.join(os.environ["USERPROFILE"], "OneDrive - Assembly", "Desktop")
    if not os.path.exists(base):
        base = os.path.join(os.environ["USERPROFILE"], "OneDrive – Assembly", "Desktop")

    pats = [
        os.path.join(base, "Report Builder Pivot (*).xlsx"),
        os.path.join(base, "Report Builder Pivot*.xlsx"),
        os.path.join(base, "stage_report.xlsx"),
    ]
    cand = []
    for p in pats:
        cand.extend(glob.glob(p))
    if not cand:
        raise FileNotFoundError(f"No Excel file found on Desktop.\nChecked: {base}")
    cand.sort(key=os.path.getmtime, reverse=True)
    print(f"[INFO] Using latest file: {cand[0]}")
    return cand[0]

def fmt_pct(x): return "0.00%" if (x is None or np.isnan(x)) else f"{x:.2f}%"
def fmt_money(x): return "$0.00" if (x is None or np.isnan(x)) else (f"${x:,.0f}" if abs(x) >= 100 else f"${x:,.2f}")
def fmt_num(x): return "0" if (x is None or np.isnan(x)) else f"{x:,.0f}"

def lob_from_campaign(x):
    s = str(x).upper()
    if "MDCR" in s: return "MDCR"
    if "CSBD" in s: return "CSBD"
    if "MDCD" in s: return "MDCD"
    return "OTHER"

def kpi(imp, clk, spd, conv):
    ctr = (clk / imp * 100) if imp > 0 else 0
    cpm = (spd / imp * 1000) if imp > 0 else 0
    return dict(ctr=ctr, cpm=cpm)

def main():
    fp = ensure_file_path(FILE_PATH)
    df = pd.read_excel(fp, sheet_name=SHEET_NAME, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    c_a, c_b, c_c = COLS["dim_a"], COLS["dim_b"], COLS["campaign"]

    # ✅ Forward fill Dimension 1, 2, 3
    df[c_a] = df[c_a].ffill()
    df[c_b] = df[c_b].ffill()
    df[c_c] = df[c_c].ffill()

    # ✅ Remove Social rows and reset index
    df = df[~df[c_b].astype(str).str.lower().eq("social")].reset_index(drop=True)

    # ✅ Convert numeric columns
    for c in [COLS["impressions"], COLS["clicks"], COLS["spend"], COLS["conversions"]]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # ✅ Add LOB + Platform mapping
    df["LOB"] = df[c_c].apply(lob_from_campaign)
    df["Platform"] = df["LOB"].map(lambda x: "Yahoo" if x == "MDCD" else ("TTD" if x in ("MDCR","CSBD") else "Other"))

    # --- KPI Aggregation
    def agg(b):
        imp = b[COLS["impressions"]].sum()
        clk = b[COLS["clicks"]].sum()
        spd = b[COLS["spend"]].sum()
        conv = b[COLS["conversions"]].sum()
        d = dict(imp=imp, clk=clk, spend=spd, conv=conv)
        d.update(kpi(imp, clk, spd, conv))
        return d

    k_all   = agg(df)
    k_ttd   = agg(df[df["Platform"]=="TTD"])
    k_yahoo = agg(df[df["Platform"]=="Yahoo"])
    mdcr    = agg(df[df["LOB"]=="MDCR"])
    csbd    = agg(df[df["LOB"]=="CSBD"])

    # --- MDCD geo leaders (for CTR & conversions already in your text)
    lead_geo_clicks = df[df["LOB"]=="MDCD"].groupby(c_a)[COLS["clicks"]].sum().sort_values(ascending=False)
    top_geo = lead_geo_clicks.index[0] if not lead_geo_clicks.empty else "NA"
    top_conv = int(df[df[c_a]==top_geo][COLS["conversions"]].sum())

    # --- NEW: MDCD top-spend geo line
    mdcd_top_spend_line = None
    mdcd_df = df[df["LOB"] == "MDCD"]
    if not mdcd_df.empty:
        spend_by_geo = (
            mdcd_df.groupby(mdcd_df[c_a].astype(str).str.strip())[COLS["spend"]]
                   .sum()
                   .sort_values(ascending=False)
        )
        if not spend_by_geo.empty:
            lead_geo_spend = spend_by_geo.index[0]
            lead_amt = spend_by_geo.iloc[0]
            if len(spend_by_geo) > 1:
                runner_geo = spend_by_geo.index[1]
                runner_amt = spend_by_geo.iloc[1]
                mdcd_top_spend_line = (
                    f"In terms of spend for the ongoing month, {lead_geo_spend} leads all MDCD geos with "
                    f"{fmt_money(lead_amt)} spent, followed closely by {runner_geo} with {fmt_money(runner_amt)}."
                )
            else:
                mdcd_top_spend_line = (
                    f"In terms of spend for the ongoing month, {lead_geo_spend} is the top MDCD geo with "
                    f"{fmt_money(lead_amt)} spent."
                )

    today = datetime.now().strftime("%b %d")

    # --- Insight Paragraphs
    # (kept exactly in your style; only added the mdcd_top_spend_line in the MDCD section)
    mdcd_spend_sentence = f"\n\n{mdcd_top_spend_line}" if mdcd_top_spend_line else ""

    insights = f"""Sydney Registration Insights  {today}

Overall Performance 

So far, Oct performance shows an overall CTR of {fmt_pct(k_all['ctr'])} with a significant difference between platforms. Yahoo reached a {fmt_pct(k_yahoo['ctr'])} CTR while TTD has a CTR of {fmt_pct(k_ttd['ctr'])}. 

Yahoo is currently exceeding TTD significantly in terms of efficiency performance but delivery through TTD is approximately 12% higher. 

ASM recommends refreshing the lists to keep the user audience fresh. In previous years when the Sydney campaign was run under the MDCD LOB, CRMs were typically updated monthly. 

Overall CPM is continuing to show good improvement from the full flight figure of $2.74, decreased to {fmt_money(k_all['cpm'])} in Oct. 

The Trade Desk 

Looking at performance by LOB, MDCR is seeing a CTR of {fmt_pct(mdcr['ctr'])} while CSBD has a CTR of {fmt_pct(csbd['ctr'])}. 

While these CTRs are below benchmark, they are not unexpected given the granularity of targeting between CRMs, Holdouts/Regular, and Registered/Unregistered users.  

Oct is going strong with great efficiency in the CSBD and MDCR campaigns, seeing CPM figures of less than {fmt_money(csbd['cpm'])} for CSBD and {fmt_money(mdcr['cpm'])} for MDCR campaigns; around 60% lower than the MDCD line. 

CARE ABCBS continues to be the top driver of conversions with 429 conversions across registered MDCR and 379 over non-registered.  

While accounting for a small portion of the overall budget, Native lines are seeing strong performance. 

So far in Oct, Native ads are seeing a CPM 7% lower than the Display counterparts. 

As of 7/14, we have moved from a strict monthly budget for Native to auto allocating budgets between the channels. So now the platform can auto-optimize and deliver where it sees the most efficiencies and conversions. 

MDCD 

Of all the states that were able to hit significant levels of spend in the first half of the month, {top_geo} led all geos with a CTR of {fmt_pct(k_yahoo['ctr'])}.{mdcd_spend_sentence}

Currently we have achieved {fmt_num(k_yahoo['conv'])} new conversions across all MDCD lines and audiences; being led {top_geo} with {fmt_num(top_conv)} new conversions.

A conversion is the combined number of landing page visits, clicks on the submit button on the state pages, and "text me a link" button clicks."""

    # --- Write Output Sheets
    wb = load_workbook(fp)
    for s in [CLEANED_SHEET, INSIGHTS_SHEET]:
        if s in wb.sheetnames:
            del wb[s]

    # Cleaned sheet
    ws_clean = wb.create_sheet(CLEANED_SHEET)
    for j, col in enumerate(df.columns, 1):
        ws_clean.cell(row=1, column=j, value=col)
    for i, row in enumerate(df.itertuples(index=False), 2):
        for j, val in enumerate(row, 1):
            ws_clean.cell(row=i, column=j, value=val)

    # Insights sheet
    ws = wb.create_sheet(INSIGHTS_SHEET)
    for i, line in enumerate(insights.splitlines(), 1):
        ws.cell(row=i, column=1, value=line)

    wb.save(fp)
    print(f"[INFO] Cleaned + Insights sheets written to {fp}")

if __name__ == "__main__":
    main()
